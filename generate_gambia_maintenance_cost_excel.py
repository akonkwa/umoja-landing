from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path("/Users/akonkwamubagwa/Documents/Playground")
OUTPUT = BASE_DIR / "EFSTH_Gambia_Maintenance_Cost_Breakdown_Annual_20260402.xlsx"

RED = "D62828"
GREY = "4A4A4A"
LIGHT_GREY = "F4F4F4"
TEXT = "111111"
WHITE = "FFFFFF"
LINE = "D9D9D9"


def money_fmt(cell):
    cell.number_format = '$#,##0.00'


def style_cell(cell, *, bold=False, fill=None, color=TEXT, align="left", size=10):
    cell.font = Font(name="Helvetica", size=size, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=True,
    )
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    thin = Side(style="thin", color=LINE)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_sheet_title(ws, title, subtitle):
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    style_cell(ws["A1"], bold=True, fill=WHITE, size=16)
    ws.merge_cells("A2:F2")
    ws["A2"] = subtitle
    style_cell(ws["A2"], color="555555")


def set_widths(ws, widths):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_summary_sheet(wb):
    ws = wb.active
    ws.title = "Summary"
    add_sheet_title(
        ws,
        "EFSTH Banjul, Gambia Maintenance Cost Breakdown",
        "Illustrative annual cost model showing why a fully burdened expert-maintenance service package for a 170 kW + 622 kWh hospital system can price at USD 18,000 per year.",
    )

    ws["A4"] = "System"
    ws["B4"] = "170 kW solar PV + 622 kWh battery energy storage system"
    ws["A5"] = "Site"
    ws["B5"] = "Edward Francis Small Teaching Hospital (EFSTH), Banjul, The Gambia"
    ws["A6"] = "Commercial framing"
    ws["B6"] = "This is a full service-maintenance retainer, not only one technician's take-home salary."
    ws["A7"] = "Annual target"
    ws["B7"] = "=Breakdown!F16"
    ws["A8"] = "Monthly equivalent"
    ws["B8"] = "=B7/12"
    ws["A9"] = "Annual cost per kW"
    ws["B9"] = "=B7/170"
    ws["A10"] = "Annual cost per kWh of storage"
    ws["B10"] = "=B7/622"

    for row in range(4, 11):
        style_cell(ws[f"A{row}"], bold=True, fill=LIGHT_GREY)
        style_cell(ws[f"B{row}"])

    money_fmt(ws["B7"])
    money_fmt(ws["B8"])
    ws["B9"].number_format = '$#,##0.00'
    ws["B10"].number_format = '$#,##0.00'

    ws["A12"] = "Why the cost is higher than a simple local salary"
    style_cell(ws["A12"], bold=True, fill=GREY, color=WHITE)
    ws.merge_cells("A12:F12")

    headers = ["Cost driver", "Why it matters at EFSTH", "Included in yearly fee"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=13, column=col)
        cell.value = header
        style_cell(cell, bold=True, fill=GREY, color=WHITE)

    rows = [
        [
            "24/7 uptime expectation",
            "Hospital power quality directly affects oxygen production and patient outcomes, so the contract carries standby and response obligations.",
            "Yes",
        ],
        [
            "Layered staffing model",
            "The slide shows local technician coverage, remote monitoring, and regional field-engineer intervention rather than one low-cost caretaker alone.",
            "Yes",
        ],
        [
            "Preventive + corrective scope",
            "Routine inspection, cleaning, thermal checks, electrical testing, firmware support, and triggered interventions all need budget cover.",
            "Yes",
        ],
        [
            "Consumables and minor spares",
            "Filters, brushes, grease, galvanizing paint, cable accessories, and broken connectors are recurring maintenance items.",
            "Yes",
        ],
        [
            "Monitoring connectivity",
            "Satellite internet / remote data service is needed to support daily monitoring and fault flagging.",
            "Yes",
        ],
    ]

    for row_idx, row in enumerate(rows, start=14):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_cell(cell)

    ws.merge_cells("A20:F20")
    ws["A20"] = "See the Breakdown sheet for line-item formulas and the Scope Link sheet for how each maintenance activity drives staffing and budget."
    style_cell(ws["A20"], fill=LIGHT_GREY)

    set_widths(ws, {1: 24, 2: 70, 3: 18, 4: 2, 5: 2, 6: 2})


def build_breakdown_sheet(wb):
    ws = wb.create_sheet("Breakdown")
    add_sheet_title(
        ws,
        "Annual Cost Model",
        "All figures in USD. Orange-highlighted rows are annual-retainer components that sum to the 18,000 target.",
    )

    headers = [
        "Category",
        "Line item",
        "Basis",
        "Rate / Unit",
        "Qty / Yr.",
        "Annual Cost",
        "Notes",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        style_cell(cell, bold=True, fill=GREY, color=WHITE)

    rows = [
        ["Labor", "Dedicated Winko maintenance lead", "Part-time retained expert allocation", 4800, 1, "=D5*E5", "Dedicated expert ownership of planning, weekly QA, reporting, and coordination."],
        ["Labor", "Local certified technician support", "Bi-weekly routine visits + callouts", 2400, 1, "=D6*E6", "Matches the maintenance plan's local contracted technician scope for inspection and cleaning."],
        ["Remote Ops", "24/7 remote monitoring and fault triage", "Annual service retainer", 1800, 1, "=D7*E7", "Daily system monitoring, alert review, and first-line troubleshooting."],
        ["Engineering", "Regional field engineer reserve", "Quarterly and triggered interventions amortized annually", 1800, 1, "=D8*E8", "Covers periodic advanced diagnostics and technical escalation from the West Africa region."],
        ["Engineering", "Annual preventive maintenance reserve", "Annual deep maintenance event", 1200, 1, "=D9*E9", "Thermal imaging, grounding tests, electrical audit, firmware updates, and protection-device testing."],
        ["Materials", "Maintenance consumables and minor spares", "Filters, brushes, grease, paint, connectors", 2100, 1, "=D10*E10", "Includes expected recurring maintenance expenditures listed by the user."],
        ["Connectivity", "Satellite internet / data service", "Annual connectivity subscription", 600, 1, "=D11*E11", "Supports the remote monitoring and alerting stack."],
        ["Logistics", "Transport, site access, and HSE overhead", "Vehicle, fuel, permits, PPE, scheduling", 900, 1, "=D12*E12", "Required to maintain regular hospital access and safe execution."],
        ["Risk", "Emergency response and uptime standby reserve", "Contract risk premium", 1200, 1, "=D13*E13", "Buffers unplanned interventions so the hospital does not wait for approvals during outages."],
        ["Admin", "Management, reporting, and contract administration", "Back-office support", 1200, 1, "=D14*E14", "Service coordination, compliance reporting, invoices, procurement handling, and QA review."],
    ]

    start_row = 5
    for row_idx, row in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_cell(cell)
            if col_idx in (4, 6):
                money_fmt(cell)

    ws["E15"] = "Total"
    style_cell(ws["E15"], bold=True, fill=GREY, color=WHITE, align="right")
    ws["F15"] = "=SUM(F5:F14)"
    style_cell(ws["F15"], bold=True, fill="FDE9D9")
    money_fmt(ws["F15"])

    ws["E16"] = "Target"
    style_cell(ws["E16"], bold=True, fill=GREY, color=WHITE, align="right")
    ws["F16"] = 18000
    style_cell(ws["F16"], bold=True, fill="FDE9D9")
    money_fmt(ws["F16"])

    ws["E17"] = "Variance"
    style_cell(ws["E17"], bold=True, fill=GREY, color=WHITE, align="right")
    ws["F17"] = "=F15-F16"
    style_cell(ws["F17"], bold=True, fill="FDE9D9")
    money_fmt(ws["F17"])

    ws["A19"] = "Interpretation"
    style_cell(ws["A19"], bold=True, fill=GREY, color=WHITE)
    ws.merge_cells("A19:G19")
    ws.merge_cells("A20:G22")
    ws["A20"] = (
        "The USD 18,000/year figure is defendable when presented as a hospital-grade maintenance package: "
        "one dedicated expert lead plus local technician support, remote monitoring, periodic regional engineer intervention, "
        "connectivity, consumables, logistics, and a response reserve. If the client wants only a basic local cleaner-technician, "
        "the number would be much lower, but that would not match the uptime promise described in the maintenance plan."
    )
    style_cell(ws["A20"])

    set_widths(ws, {1: 14, 2: 32, 3: 34, 4: 13, 5: 12, 6: 14, 7: 44})


def build_scope_link_sheet(wb):
    ws = wb.create_sheet("Scope Link")
    add_sheet_title(
        ws,
        "Scope-to-Cost Linkage",
        "Maps the maintenance-plan activities to the resource buckets used in the annual cost model.",
    )

    headers = ["Maintenance activity", "Frequency", "Responsible party", "Budget bucket", "Why it drives cost"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        style_cell(cell, bold=True, fill=GREY, color=WHITE)

    rows = [
        ["Remote system monitoring (energy output, inverter status, battery SOC and temperature)", "Daily", "Winko Solar (remote)", "24/7 remote monitoring and fault triage", "Requires staffed monitoring tools, dashboards, alert management, and daily review."],
        ["Automated fault alerting and performance flagging", "Continuous", "Winko Solar (remote)", "24/7 remote monitoring and fault triage", "Continuous anomaly detection implies a paid service layer, not a casual periodic check."],
        ["Visual site inspection (panels, mounting, cables, connectors, corrosion)", "Bi-weekly", "Local certified technician", "Dedicated Winko maintenance lead + Local certified technician support", "Regular site attendance needs a qualified local person and supervision."],
        ["Panel cleaning based on observed soiling", "As needed / bi-weekly", "Local certified technician", "Local certified technician support + Maintenance consumables and minor spares", "Cleaning materials and labor recur throughout the year."],
        ["Technical inspection (string measurements, inverter logs, battery performance, tightening, device testing)", "Quarterly", "Winko Solar field engineer", "Regional field engineer reserve", "Higher-skill visits are less frequent but materially raise the monthly retainer when amortized."],
        ["Preventive maintenance (thermal imaging, grounding test, electrical audit, firmware updates)", "Annual", "Winko Solar field engineer", "Annual preventive maintenance reserve", "Specialized tools and engineering time are usually budgeted as an annual event spread across months."],
        ["Performance-based intervention for drops of ~10% or more", "Triggered", "Winko Solar field engineer", "Emergency response and uptime standby reserve", "The service provider must hold capacity for unscheduled response."],
    ]

    for row_idx, row in enumerate(rows, start=5):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_cell(cell)

    set_widths(ws, {1: 46, 2: 18, 3: 28, 4: 34, 5: 50})


def build_assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions")
    add_sheet_title(
        ws,
        "Assumptions and Messaging Notes",
        "Use these notes when presenting the spreadsheet to explain what the 18,000/year covers.",
    )

    entries = [
        ("1. Pricing basis", "The workbook assumes the 18,000 figure is in USD per year."),
        ("2. What is being priced", "A hospital-critical maintenance service package, not only one technician's wage."),
        ("3. Site context", "EFSTH is presented as a mission-critical health facility where downtime has a high operational and reputational cost."),
        ("4. Staffing logic", "The annual price reflects a layered service model: dedicated maintenance lead, local technician support, remote operations, and periodic field engineer involvement."),
        ("5. User-provided line items", "Dedicated permanent Winko employee, weekly inspection and cleaning on schedule, maintenance expenditures, and satellite internet were all included explicitly."),
        ("6. Commercial caution", "If a buyer interprets the figure as the cash salary for one Gambian technician, they may reject it. Present it as a bundled O&M retainer with service guarantees and risk coverage."),
        ("7. Adjustment guidance", "If you want a leaner version, the easiest reductions are the standby reserve, admin overhead, and field-engineer reserve."),
    ]

    ws["A4"] = "Item"
    ws["B4"] = "Note"
    style_cell(ws["A4"], bold=True, fill=GREY, color=WHITE)
    style_cell(ws["B4"], bold=True, fill=GREY, color=WHITE)

    for row_idx, (item, note) in enumerate(entries, start=5):
        ws[f"A{row_idx}"] = item
        ws[f"B{row_idx}"] = note
        style_cell(ws[f"A{row_idx}"], bold=True, fill=LIGHT_GREY)
        style_cell(ws[f"B{row_idx}"])

    set_widths(ws, {1: 24, 2: 96})


def main():
    wb = Workbook()
    build_summary_sheet(wb)
    build_breakdown_sheet(wb)
    build_scope_link_sheet(wb)
    build_assumptions_sheet(wb)

    for ws in wb.worksheets:
        ws.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
