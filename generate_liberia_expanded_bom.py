from pathlib import Path

from openpyxl import load_workbook


TEMPLATE = Path("/Users/akonkwamubagwa/Documents/Winko Expanded BoM 20260217.xlsx")
OUTPUT = Path("/Users/akonkwamubagwa/Documents/Winko Expanded BoM Liberia 20260317.xlsx")


def setv(ws, cell, value):
    ws[cell] = value


wb = load_workbook(TEMPLATE)
ws = wb["Expanded_BoM"]
ws2 = wb["Sheet1"]
for name in list(wb.sheetnames):
    if name not in {"Expanded_BoM", "Sheet1"}:
        wb.remove(wb[name])

# Expanded_BoM sheet
setv(ws, "B3", "PV Panels")
setv(ws, "C3", "Solar PV Modules (90 units)")
setv(ws, "D3", "High-efficiency photovoltaic modules (465 Wp class)")
setv(ws, "E3", 16041.92)

setv(ws, "B4", "PV Connectors")
setv(ws, "C4", "MC4 Connectors & DC Connectors")
setv(ws, "D4", "Industry-standard MC4-compatible connectors and DC terminations")
setv(ws, "E4", 759.38)

setv(ws, "B5", "PV Cabling")
setv(ws, "C5", "DC String Cabling")
setv(ws, "D5", "UV-resistant solar DC cabling including routing, terminations, and protection for reliable long-term operation in outdoor conditions.")
setv(ws, "E5", 1198.70)

setv(ws, "B6", "PV Modules Total")
setv(ws, "C6", "Solar Panels")
setv(ws, "D6", "90 modules, including MC4 connectors and DC wiring")
setv(ws, "E6", "=SUM(E3:E5)")

setv(ws, "B7", "BESS Total")
setv(ws, "C7", "Battery Modules")
setv(ws, "D7", "LiFePO4 battery cells and racks")
setv(ws, "E7", 20733.88)

setv(ws, "B8", "BESS Total")
setv(ws, "C8", "Battery Management System")
setv(ws, "D8", "BMS, enclosure, monitoring interface")
setv(ws, "E8", 9266.12)

setv(ws, "B9", "Battery Energy Storage System (BESS Total")
setv(ws, "C9", None)
setv(ws, "D9", None)
setv(ws, "E9", "=SUM(E7:E8)")

setv(ws, "B11", "Inverter")
setv(ws, "C11", "Hybrid Inverter Units")
setv(ws, "D11", "125 kW class modular hybrid inverter with communication modules")
setv(ws, "E11", 12000)

setv(ws, "B13", "Mounting")
setv(ws, "C13", "Racking System")
setv(ws, "D13", "Rails, clamps, rooftop structural attachments")
setv(ws, "E13", 5490.20)

setv(ws, "B14", "Mounting")
setv(ws, "C14", "DC Combiner & Cabling")
setv(ws, "D14", "Combiner boxes and DC cabling")
setv(ws, "E14", 2509.80)

setv(ws, "B15", "Mounting Total")
setv(ws, "C15", None)
setv(ws, "D15", None)
setv(ws, "E15", "=SUM(E13:E14)")

setv(ws, "B17", "Electrical BOS")
setv(ws, "C17", "ATS")
setv(ws, "D17", "Automatic Transfer Switch")
setv(ws, "E17", 1963.90)

setv(ws, "B18", "Electrical BOS")
setv(ws, "C18", "Breakers (MCCB)")
setv(ws, "D18", "Main protection breakers")
setv(ws, "E18", 1472.93)

setv(ws, "B19", "Electrical BOS")
setv(ws, "C19", "Surge Protection")
setv(ws, "D19", "SPD devices")
setv(ws, "E19", 981.95)

setv(ws, "B20", "Electrical BOS")
setv(ws, "C20", "AC Cabling")
setv(ws, "D20", "AC wiring and terminations")
setv(ws, "E20", 1599.27)

setv(ws, "B21", "Electrical BOS")
setv(ws, "C21", "Grounding System")
setv(ws, "D21", "Earthing and grounding infrastructure")
setv(ws, "E21", 981.95)

setv(ws, "B22", "Electrical BOS Total ")
setv(ws, "C22", None)
setv(ws, "D22", None)
setv(ws, "E22", "=SUM(E17:E21)")

setv(ws, "B24", "Logistics")
setv(ws, "C24", "Import & Shipping")
setv(ws, "D24", "Freight, port handling, customs")
setv(ws, "E24", 8000)

setv(ws, "B25", "Installation")
setv(ws, "C25", "On-site Works")
setv(ws, "D25", "Labor and installation")
setv(ws, "E25", 22300)

setv(ws, "B26", "Engineering")
setv(ws, "C26", "Design & Commissioning")
setv(ws, "D26", "Engineering, travel, commissioning")
setv(ws, "E26", 15400)

setv(ws, "B28", "Logistics")
setv(ws, "C28", "International Freight")
setv(ws, "D28", "Containerized shipping of equipment")
setv(ws, "E28", 3000)

setv(ws, "B29", "Logistics")
setv(ws, "C29", "Port Handling & Customs")
setv(ws, "D29", "Clearance, duties, port fees")
setv(ws, "E29", 2500)

setv(ws, "B30", "Logistics")
setv(ws, "C30", "Inland Transport")
setv(ws, "D30", "Transport from port to site")
setv(ws, "E30", 2500)

setv(ws, "B31", "Logistics Total")
setv(ws, "C31", "Import & Shipping")
setv(ws, "D31", "Freight, port handling, customs")
setv(ws, "E31", "=SUM(E28:E30)")

setv(ws, "B33", "Installation")
setv(ws, "C33", "Civil Works")
setv(ws, "D33", "Mounting structures, trenching, site prep")
setv(ws, "E33", 7212.94)

setv(ws, "B34", "Installation")
setv(ws, "C34", "Electrical Installation")
setv(ws, "D34", "DC/AC wiring, equipment installation")
setv(ws, "E34", 9016.17)

setv(ws, "B35", "Installation")
setv(ws, "C35", "Labor & Supervision")
setv(ws, "D35", "Skilled labor, on-site supervision")
setv(ws, "E35", 6070.89)

setv(ws, "B36", "Installation Total")
setv(ws, "C36", "On-site Works")
setv(ws, "D36", "Labor and installation")
setv(ws, "E36", "=SUM(E33:E35)")

setv(ws, "B38", "Engineering")
setv(ws, "C38", "System Design")
setv(ws, "D38", "Technical design, load analysis, drawings")
setv(ws, "E38", 5310.34)

setv(ws, "B39", "Engineering")
setv(ws, "C39", "Project Engineering")
setv(ws, "D39", "Validation, system optimization")
setv(ws, "E39", 3540.23)

setv(ws, "B40", "Engineering")
setv(ws, "C40", "Commissioning & Testing")
setv(ws, "D40", "System testing, commissioning")
setv(ws, "E40", 3540.23)

setv(ws, "B41", "Engineering")
setv(ws, "C41", "Travel & Site Visits")
setv(ws, "D41", "Engineering travel and logistics")
setv(ws, "E41", 3009.20)

setv(ws, "B42", "Engineering Total")
setv(ws, "C42", "Design & Commissioning")
setv(ws, "D42", "Engineering, travel, commissioning")
setv(ws, "E42", "=SUM(E38:E41)")

setv(ws, "B44", "EPC- Project Delivery & Performance Assurance")
setv(ws, "C44", "Project Management")
setv(ws, "D44", "Timeline, coordination, reporting")
setv(ws, "E44", 12672.42)

setv(ws, "B45", "EPC- Project Delivery & Performance Assurance")
setv(ws, "C45", "Procurement Coordination")
setv(ws, "D45", "Supplier management, logistics coordination")
setv(ws, "E45", 9504.31)

setv(ws, "B46", "EPC- Project Delivery & Performance Assurance")
setv(ws, "C46", "Quality Control")
setv(ws, "D46", "Installation oversight, QA/QC processes")
setv(ws, "E46", 7920.26)

setv(ws, "B47", "EPC- Project Delivery & Performance Assurance")
setv(ws, "C47", "Risk Mitigation")
setv(ws, "D47", "Contingency planning, issue resolution")
setv(ws, "E47", 11088.36)

setv(ws, "B48", "EPC- Project Delivery & Performance Assurance")
setv(ws, "C48", "Performance Assurance")
setv(ws, "D48", "System reliability, uptime accountability")
setv(ws, "E48", 18814.65)

setv(ws, "B49", "EPC - Project Delivery & Performance Assurance Total")
setv(ws, "C49", "Project Delivery")
setv(ws, "D49", "Project management and performance assurance")
setv(ws, "E49", "=SUM(E44:E48)")

setv(ws, "B51", "Automatic Cleaning System")
setv(ws, "C51", "Zone A: Rail system ")
setv(ws, "D51", "Rail-based cleaning system for one half of the array, including carriage, brush, motor, and controls")
setv(ws, "E51", 8500)

setv(ws, "B52", "Automatic Cleaning System")
setv(ws, "C52", "Zone B: Tethered robot ")
setv(ws, "D52", "Tethered cleaning robot with gyroscope stabilization, tether reel, and docking setup")
setv(ws, "E52", 11000)

setv(ws, "B53", "Automatic Cleaning System")
setv(ws, "C53", "Water infrastructure ")
setv(ws, "D53", "Header tank, hoses, fittings, and metering for cleaning operations")
setv(ws, "E53", 2000)

setv(ws, "B54", "Automatic Cleaning System")
setv(ws, "C54", "Instrumentation and monitoring integration")
setv(ws, "D54", "Cleaning cycle logging, system controls, and monitoring integration")
setv(ws, "E54", 2500)

setv(ws, "B55", "Automatic Cleaning System")
setv(ws, "C55", "Installation & system integration (incremental to core EPC scope)")
setv(ws, "D55", "Installation labor, integration, rigging, and calibration")
setv(ws, "E55", 3000)

setv(ws, "B56", "Automatic Cleaning System")
setv(ws, "C56", "System Engineering, Commissioning & 6-month performance monitoring")
setv(ws, "D56", "Engineering, commissioning, and monitored pilot performance evaluation")
setv(ws, "E56", 3000)

setv(ws, "B57", "Automatic Cleaning System")
setv(ws, "C57", None)
setv(ws, "D57", None)
setv(ws, "E57", None)

setv(ws, "B58", "Automatic Cleaning System")
setv(ws, "C58", "Shipping ")
setv(ws, "D58", "Shipping allowance and packaging consolidated with main shipment")
setv(ws, "E58", 3000)

setv(ws, "B59", "Automatic Cleaning System Total")
setv(ws, "C59", "Automated Cleaning System")
setv(ws, "D59", "Rail + robotic cleaning system")
setv(ws, "E59", "=SUM(E51:E58)")

setv(ws, "E60", "=E59+E49+E42+E36+E31+E22+E15+E9+E6+E11")

# Sheet1 cleaning breakout
setv(ws2, "C3", "Zone A: Rail system ")
setv(ws2, "D3", "Rail-based cleaning system for one half of the array, including carriage, brush, motor, and controls")
setv(ws2, "E3", 8500)
setv(ws2, "C4", "Zone B: Tethered robot ")
setv(ws2, "D4", "Tethered cleaning robot with gyroscope stabilization, tether reel, and docking setup")
setv(ws2, "E4", 11000)
setv(ws2, "C5", "Water infrastructure ")
setv(ws2, "D5", "Header tank, hoses, fittings, and meter")
setv(ws2, "E5", 2000)
setv(ws2, "C6", "Instrumentation & monitoring integration ")
setv(ws2, "D6", "Cleaning controls, logging, and monitoring integration")
setv(ws2, "E6", 2500)
setv(ws2, "C7", "Installation labor (incremental) ")
setv(ws2, "D7", "Installation labor, rigging, and calibration")
setv(ws2, "E7", 3000)
setv(ws2, "C8", "Engineering, commissioning, 6-month monitoring ")
setv(ws2, "D8", "Engineering, commissioning, and monitored pilot performance evaluation")
setv(ws2, "E8", 3000)
setv(ws2, "C9", "Shipping ")
setv(ws2, "D9", "Shipping allowance + packaging consolidated with main shipment")
setv(ws2, "E9", 3000)
setv(ws2, "C10", None)
setv(ws2, "D10", None)
setv(ws2, "E10", None)
setv(ws2, "E11", "=SUM(E3:E10)")

wb.save(OUTPUT)
print(OUTPUT)
